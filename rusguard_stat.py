import urllib3
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.filters import AutoFilter, CustomFilter
from requests import Session
from zeep import Client
from zeep.transports import Transport
import ssl
import datetime
import ast

#####  disable ssl warning
urllib3.disable_warnings()
ssl._create_default_https_context = ssl._create_unverified_context

##### connect to rusguard server
from zeep.wsse.username import UsernameToken
session = Session()
session.verify =False
transport=Transport(session=session)

client = Client('',
                wsse=UsernameToken("", ""), transport=transport)


# Создание нового документа Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Проходы"

# Установка шрифта
font = Font(name='Times New Roman', size=14)

# Заголовки
headers = ["ФИО", "Группа", "Дата", "Первый вход", "Последний выход", "Промежуточные входы", "Промежуточные выходы"]
ws.append(headers)

# Применение стиля к заголовкам
for row in ws.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.font = Font(name='Times New Roman', size=16, bold=True)

# Общий список для всех записей
all_log_messages = []

# Двери и типы событий
doors = [
]

# Получение записей для каждой двери и типа события
for door_ID, types_m in doors:
    events = client.service.GetEventsByDeviceIDs(0, datetime.datetime(2023, 5, 10, 00, 1), datetime.datetime(2023, 5, 15, 00, 00), msgTypes="Information", msgSubTypes=types_m, deviceIDs=[door_ID])
    log_messages = events['Messages']['LogMessage']
    all_log_messages.extend(log_messages)

# Определенные ключи, по которым вы хотите извлекать значения
desired_keys = ['EmployeeFirstName', 'EmployeeSecondName', 'EmployeeSecondName', 'EmployeeGroupName', 'Message', 'DateTime']

# Словарь для группировки списков по ФИО
grouped_logs = {}

# Группировка записей
for log in all_log_messages:
    if log['EmployeeLastName'] is not None and log['EmployeeFirstName'] is not None and log['EmployeeSecondName'] is not None:
        key = ' '.join([log['EmployeeLastName'], log['EmployeeFirstName'], log['EmployeeSecondName']])
        if key not in grouped_logs:
            grouped_logs[key] = {'group': log['EmployeeGroupName'], 'events': []}
        formatted_datetime = log['DateTime'].strftime('%Y-%m-%d %H:%M:%S')
        event_type = log['Message']
        grouped_logs[key]['events'].append((formatted_datetime, event_type))



for key, values in grouped_logs.items():
    fio = key
    group = values['group']
    date = values['events'][0][0].split(' ')[0]  # Используем первую дату из событий

    entry_times = [entry[0].split(' ')[-1] for entry in values['events'] if entry[1] == "Вход"]
    exit_times = [entry[0].split(' ')[-1] for entry in values['events'] if entry[1] == "Выход"]

    first_entry = ""
    last_exit = ""
    
    # Удаляем первый вход и последний выход из промежуточных входов и выходов
    intermediate_entries = "\n".join(sorted(entry_times))  # Сортировка по возрастанию
    if entry_times:
        first_entry = intermediate_entries.split('\n')[0]
        intermediate_entries = "\n".join(entry_times[1:])
    
    intermediate_exits = "\n".join(sorted(exit_times))  # Сортировка по возрастанию
    if exit_times:
        last_exit = intermediate_exits.split('\n')[-1]
        intermediate_exits = "\n".join(exit_times[:-1])


    row = [fio, group, date, first_entry, last_exit, intermediate_entries, intermediate_exits]
    ws.append(row)


# Автоматическое настройка выравнивания текста по центру
for column_cells in ws.columns:
    for cell in column_cells:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


#Автоматическая настройка ширины столбцов по содержимому
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            lines = cell.value.split("\n")
            max_line_length = max(len(line) for line in lines)
            if max_line_length > max_length:
                max_length = max_line_length
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Добавляем небольшой запас
    ws.column_dimensions[column_letter].width = adjusted_width


# Настройка высоты строк с учетом содержимого
for row in ws.iter_rows(min_row=2):  # Начиная со второй строки (после заголовков)
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical='center')  # Выравнивание по середине


# Добавление границ для всей таблицы
border = openpyxl.styles.Border(
    left=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
    top=openpyxl.styles.Side(style='thin'),
    bottom=openpyxl.styles.Side(style='thin')
)
for row in ws.iter_rows():
    for cell in row:
        cell.border = border


# Добавление фильтров к столбцам Группа и Дата
ws.auto_filter.ref = "B1:C2000"  # Индекс 1 - столбец "Группа", Индекс 2 - столбец "Дата"


# Сохранение файла
wb.save("log_data.xlsx")












                        #####           Запасной, рабочий вариант

# #####  disable ssl warning
# urllib3.disable_warnings()
# ssl._create_default_https_context = ssl._create_unverified_context

# ##### connect to rusguard server
# from zeep.wsse.username import UsernameToken
# session = Session()
# session.verify =False
# transport=Transport(session=session)

# client = Client('',
#                 wsse=UsernameToken("Админ", "P@ssw0rd47"), transport=transport)


# # Создание нового документа Excel
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Проходы"

# # Установка шрифта
# font = Font(name='Times New Roman', size=14)

# # Заголовки
# headers = ["ФИО", "Группа", "Дата", "Первый вход", "Последний выход", "Промежуточные входы", "Промежуточные выходы"]
# ws.append(headers)

# # Применение стиля к заголовкам
# for row in ws.iter_rows(min_row=1, max_row=1):
#     for cell in row:
#         cell.font = Font(name='Times New Roman', size=16, bold=True)

# # Общий список для всех записей
# all_log_messages = []

# # Двери и типы событий
# doors = [

# ]

# # Получение записей для каждой двери и типа события
# for door_ID, types_m in doors:
#     events = client.service.GetEventsByDeviceIDs(0, datetime.datetime(2023, 5, 10, 00, 1), datetime.datetime(2023, 5, 15, 00, 00), msgTypes="Information", msgSubTypes=types_m, deviceIDs=[door_ID])
#     log_messages = events['Messages']['LogMessage']
#     all_log_messages.extend(log_messages)

# # Определенные ключи, по которым вы хотите извлекать значения
# desired_keys = ['EmployeeFirstName', 'EmployeeSecondName', 'EmployeeSecondName', 'EmployeeGroupName', 'Message', 'DateTime']

# # Словарь для группировки списков по ФИО
# grouped_logs = {}

# # Группировка записей
# for log in all_log_messages:
#     if log['EmployeeLastName'] is not None and log['EmployeeFirstName'] is not None and log['EmployeeSecondName'] is not None:
#         key = ' '.join([log['EmployeeLastName'], log['EmployeeFirstName'], log['EmployeeSecondName']])
#         if key not in grouped_logs:
#             grouped_logs[key] = {'group': log['EmployeeGroupName'], 'events': []}
#         formatted_datetime = log['DateTime'].strftime('%Y-%m-%d %H:%M:%S')
#         event_type = log['Message']
#         grouped_logs[key]['events'].append((formatted_datetime, event_type))



# for key, values in grouped_logs.items():
#     fio = key
#     group = values['group']
#     date = values['events'][0][0].split(' ')[0]  # Используем первую дату из событий

#     entry_times = [entry[0].split(' ')[-1] for entry in values['events'] if entry[1] == "Вход"]
#     exit_times = [entry[0].split(' ')[-1] for entry in values['events'] if entry[1] == "Выход"]

#     first_entry = ""
#     last_exit = ""
#     intermediate_entries = "\n".join(sorted(entry_times))  # Сортировка по возрастанию
#     intermediate_exits = "\n".join(sorted(exit_times))  # Сортировка по возрастанию

#     if entry_times:
#         first_entry = min(entry_times)
#     if exit_times:
#         last_exit = max(exit_times)
    
    
#     row = [fio, group, date, first_entry, last_exit, intermediate_entries, intermediate_exits]
#     ws.append(row)


# # Автоматическое настройка выравнивания текста по центру
# for column_cells in ws.columns:
#     for cell in column_cells:
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# #Автоматическая настройка ширины столбцов по содержимому
# for column in ws.columns:
#     max_length = 0
#     column_letter = column[0].column_letter
#     for cell in column:
#         try:
#             lines = cell.value.split("\n")
#             max_line_length = max(len(line) for line in lines)
#             if max_line_length > max_length:
#                 max_length = max_line_length
#         except:
#             pass
#     adjusted_width = (max_length + 2) * 1.2  # Добавляем небольшой запас
#     ws.column_dimensions[column_letter].width = adjusted_width


# # Настройка высоты строк с учетом содержимого
# for row in ws.iter_rows(min_row=2):  # Начиная со второй строки (после заголовков)
#     for cell in row:
#         cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical='center')  # Выравнивание по середине


# # Добавление границ для всей таблицы
# border = openpyxl.styles.Border(
#     left=openpyxl.styles.Side(style='thin'),
#     right=openpyxl.styles.Side(style='thin'),
#     top=openpyxl.styles.Side(style='thin'),
#     bottom=openpyxl.styles.Side(style='thin')
# )
# for row in ws.iter_rows():
#     for cell in row:
#         cell.border = border


# # Добавление фильтров к столбцам Группа и Дата
# ws.auto_filter.ref = "B1:C2000"  # Индекс 1 - столбец "Группа", Индекс 2 - столбец "Дата"


# # Сохранение файла
# wb.save("log_data.xlsx")

