import json
import requests
from openpyxl import load_workbook

# Загрузка таблицы
workbook = load_workbook("/file")

results = []  # Список для хранения результатов

day_mapping = {
    'понедельник': 'monday',
    'вторник': 'tuesday',
    'среда': 'wednesday',
    'четверг': 'thursday',
    'пятница': 'friday',
    'суббота': 'saturday',
    'воскресенье': 'sunday'
}
group = None  # Переменная для хранения предыдущего значения
# Проход по всем листам в книге
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    for column in sheet.iter_cols():
        for cell in column:
            if cell.value in day_mapping:
                day_of_week = sheet.cell(row=cell.row, column=cell.column).value
                day_of_week_eng = day_mapping.get(day_of_week)
                if day_of_week == 'понедельник':
                    group = sheet.cell(row=cell.row - 2, column=cell.column + 1 ).value

                output = {"groupName": group, "weekDay": day_of_week_eng, "pairs": []}

                # Добавление значений из соседних ячеек
                for i in range(0, 6):  # 6 строк вниз
                    pair_column = cell.column + 2
                    room_column = cell.column + 3
                    pair_number_column = cell.column + 1

                    pair_value = sheet.cell(row=cell.row + i, column=pair_column).value
                    room_value = sheet.cell(row=cell.row + i, column=room_column).value
                    pair_number_value = sheet.cell(row=cell.row + i, column=pair_number_column).value

                    if pair_value is not None:
                        pair_value = ' '.join(pair_value.split())
                        if room_value and room_value != 0:
                            room_value = ' '.join(str(room_value).split())
                            row_values = {"pairNum": pair_number_value, "pairName": pair_value, "pairCab": room_value}
                        else:
                            row_values = {"pairNum": pair_number_value, "pairName": pair_value, "pairCab": ""}
                        output["pairs"].append(row_values)

                if output["pairs"]:
                    results.append(output)


payload = {
    "password": "send-pass",
    "groups": results
}


json_payload = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)

# with open("лог.json", "w", encoding='utf-8') as data:
#      data.write(json_payload)

    # Отправка файла по ссылке
server_url = ""  # Замените на вашу ссылку
response = requests.post(server_url, json=payload, headers={'Content-Type': 'application/json'})
print(f"Отправка результатов на сайт для листа '{sheet_name}': {response.status_code} - {response.text}")
 
# Закрытие таблицы
workbook.close()